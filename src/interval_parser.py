"""Helpers for extracting hole metadata from gradation report workbooks."""

import logging
import re
from functools import lru_cache
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Regex patterns for hole ID extraction
HOLE_RES = [
    re.compile(r"\bT[-_ ]?([A-Za-z0-9]+)\b", re.I),  # T3, T-3, T 3
    re.compile(r"\bBore[-_ ]?Hole[-_ ]?([A-Za-z0-9]+)\b", re.I),
]

# Regex patterns for interval extraction
INTERVAL_RES = [
    re.compile(
        r"\b(\d+(?:\.\d+)?)\s*[_-]\s*(\d+(?:\.\d+)?)\s*(?:ft|feet|'|\")?\b",
        re.I,
    ),  # 5_10, 5-10
    re.compile(
        r"\b(\d+(?:\.\d+)?)\s*to\s*(\d+(?:\.\d+)?)\s*(?:ft|feet|'|\")?\b",
        re.I,
    ),  # 5 to 10
]


def _normalize_depth(value: float) -> float | int:
    value = float(value)
    if abs(value - round(value)) < 1e-3:
        return int(round(value))
    return round(value, 3)


def _normalize_interval_bounds(
    start: float,
    end: float,
) -> Tuple[float | int, float | int]:
    start = float(start)
    end = float(end)
    if start >= end:
        raise ValueError("Invalid interval: start >= end")
    return _normalize_depth(start), _normalize_depth(end)


def _coerce_numeric(value) -> Optional[float]:
    if value is None:
        return None
    # Avoid pandas dependency for simple NA checks
    if isinstance(value, (int, float)):
        # Filter out NaN without pandas
        try:
            if value != value:  # NaN != NaN
                return None
        except Exception:
            return None
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
        if match:
            try:
                return float(match.group(0))
            except ValueError:
                return None
    return None


@lru_cache(maxsize=16)
def _load_workbook(path: str):
    return load_workbook(path, data_only=True)


def _ensure_pandas():
    try:
        import pandas as pd  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "pandas is required to compute FM from sieve tables"
        ) from exc
    return pd


def parse_hole_id_from_title(title: str) -> Optional[str]:
    """Extract hole ID from filename or folder name.

    Args:
        title: Filename or folder name

    Returns:
        Hole ID (e.g., "T3") or None if not found
    """
    for rx in HOLE_RES:
        m = rx.search(title)
        if m:
            return m.group(1).upper()
    return None


def parse_hole_id_from_sheet(workbook) -> Optional[str]:
    """Attempt to infer hole ID from worksheet contents."""

    sh = workbook.active
    for row in sh.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str):
                candidate = parse_hole_id_from_title(value)
                if candidate:
                    return candidate
    return None


def parse_interval_from_title(title: str) -> Optional[Tuple[int, int]]:
    """Extract depth interval from filename.

    Args:
        title: Filename

    Returns:
        Tuple of (start_ft, end_ft) or None if not found

    Raises:
        ValueError: If interval is invalid (start >= end)
    """
    # Normalize various dash/quotes
    t = title.replace("'", "'").replace("–", "-").replace("—", "-")

    for rx in INTERVAL_RES:
        m = rx.search(t)
        if m:
            a, b = _normalize_interval_bounds(m.group(1), m.group(2))
            return a, b
    return None


def parse_interval_from_sheet(
    path: str,
    workbook=None,
) -> Optional[Tuple[int, int]]:
    """Extract depth interval from worksheet when title lacks it.

    Args:
        path: Path to Excel file

    Returns:
        Tuple of (start_ft, end_ft) or None if not found
    """

    depth_pattern = re.compile(
        r"(\d+(?:\.\d+)?)\s*(?:-|to)\s*(\d+(?:\.\d+)?)",
        re.IGNORECASE,
    )

    wb = workbook or _load_workbook(str(Path(path)))
    sh = wb.active

    depth_col: Optional[int] = None
    start_col: Optional[int] = None
    end_col: Optional[int] = None
    spec_col: Optional[int] = None

    for row in sh.iter_rows(values_only=True):
        if not any(row):
            continue

        # Detect header columns
        for idx, value in enumerate(row):
            if isinstance(value, str):
                lower = value.strip().lower()
                if depth_col is None and lower == "depth":
                    depth_col = idx
                if start_col is None and "start" in lower and "depth" in lower:
                    start_col = idx
                if end_col is None and "end" in lower and "depth" in lower:
                    end_col = idx
                if spec_col is None and "spec" in lower and "range" in lower:
                    spec_col = idx

        # Check for inline interval text
        for idx, value in enumerate(row):
            if isinstance(value, str):
                if spec_col is not None and idx == spec_col:
                    continue
                match = depth_pattern.search(value)
                if match:
                    return _normalize_interval_bounds(
                        match.group(1),
                        match.group(2),
                    )

        # Check cells identified as depth/start/end columns
        if start_col is not None and end_col is not None:
            start_val = row[start_col] if start_col < len(row) else None
            end_val = row[end_col] if end_col < len(row) else None
            start_num = _coerce_numeric(start_val)
            end_num = _coerce_numeric(end_val)
            if (
                start_num is not None
                and end_num is not None
                and start_num < end_num
            ):
                return _normalize_interval_bounds(start_num, end_num)

        if depth_col is not None and depth_col < len(row):
            depth_val = row[depth_col]
            if isinstance(depth_val, str):
                match = depth_pattern.search(depth_val)
                if match:
                    return _normalize_interval_bounds(
                        match.group(1),
                        match.group(2),
                    )

    return None


def parse_location_cell(raw: str) -> Tuple[float, float]:
    """Parse location cell to extract latitude and longitude.

    Supports:
    - Decimal: 32.483210, -96.361122
    - Decimal with N/W: 32.483210 N, 96.361122 W
    - DMS: 32°28'59.6" N, 96°21'40.0" W

    Args:
        raw: Raw location string from Excel cell

    Returns:
        Tuple of (latitude, longitude)

    Raises:
        ValueError: If location format is unrecognized
    """
    s = raw.strip()

    # Try decimal format first
    m = re.search(r"([-+]?\d+(?:\.\d+)?)\s*,\s*([-+]?\d+(?:\.\d+)?)", s)
    if m:
        lat, lon = float(m.group(1)), float(m.group(2))

        # Handle N/S W/E suffix flipping
        ns = re.search(r"([NS])", s, re.I)
        ew = re.search(r"([EW])", s, re.I)

        if ns and ns.group(1).upper() == "S":
            lat = -abs(lat)
        if ew and ew.group(1).upper() == "W":
            lon = -abs(lon)

        return lat, lon

    # Try DMS format
    m = re.search(
        r"(\d+)[°\s]+(\d+)[\'\s]+([\d.]+)\"?\s*([NS])\s*,?\s*"
        r"(\d+)[°\s]+(\d+)[\'\s]+([\d.]+)\"?\s*([EW])",
        s,
        re.I,
    )
    if m:

        def dms_to_dd(sign, d, mnt, sec):
            return sign * (abs(d) + mnt / 60 + sec / 3600)

        lat = dms_to_dd(
            1 if m.group(4).upper() == "N" else -1,
            int(m.group(1)),
            int(m.group(2)),
            float(m.group(3)),
        )
        lon = dms_to_dd(
            1 if m.group(8).upper() == "E" else -1,
            int(m.group(5)),
            int(m.group(6)),
            float(m.group(7)),
        )
        return lat, lon

    # Fallback: examine decimal values and apply direction hints if present
    decimal_tokens = re.findall(r"[-+]?\d+(?:\.\d+)?", s)
    if len(decimal_tokens) >= 2:
        lat = float(decimal_tokens[0])
        lon = float(decimal_tokens[1])

        ns = re.search(r"([NS])", s, re.I)
        ew = re.search(r"([EW])", s, re.I)

        if ns and ns.group(1).upper() == "S":
            lat = -abs(lat)
        if ew and ew.group(1).upper() == "W":
            lon = -abs(lon)
        return lat, lon

    msg = f"Unrecognized Location format: {raw}"
    raise ValueError(msg)


def extract_fm_from_xlsx(path: str, workbook=None) -> float:
    """Extract Fineness Modulus from Excel file.

    Tries multiple strategies:
    1. Find "Fineness Modulus" or "FM" label and extract numeric value
    2. Compute from sieve table if present (fallback)

    Args:
        path: Path to Excel file

    Returns:
        Fineness Modulus value

    Raises:
        ValueError: If FM cannot be found or computed
    """
    wb = workbook or _load_workbook(str(Path(path)))
    sh = wb.active

    # Strategy 1: Find FM label and extract value
    for row in sh.iter_rows(values_only=True):
        for c, v in enumerate(row):
            if isinstance(v, str) and (
                "fineness modulus" in v.lower() or v.strip().lower() == "fm"
            ):
                # Look right within 8 cells for numeric value
                for off in range(1, 9):
                    if c + off < len(row):
                        adjacent_val = row[c + off]
                        # Only accept numeric values (int/float), not strings that might contain numbers
                        if isinstance(adjacent_val, (int, float)):
                            # Filter out NaN
                            try:
                                if adjacent_val != adjacent_val:  # NaN != NaN
                                    continue
                            except Exception:
                                continue
                            nv = float(adjacent_val)
                            if 0.5 <= nv <= 7.0:
                                filename = Path(path).name
                                logger.debug(
                                    "Found FM value %s from label in %s",
                                    nv,
                                    filename,
                                )
                                return nv

                # Skip inline matching - too risky to extract from strings

    # Strategy 2: Compute from sieve table (fallback)
    filename = Path(path).name
    logger.warning(
        "FM label not found, attempting to compute from sieve table in %s",
        filename,
    )
    try:
        fm = compute_fm_from_sieve_table(sh, path)
        logger.info(
            "Computed FM value %s from sieve table in %s",
            fm,
            Path(path).name,
        )
        return fm
    except Exception as e:
        logger.error(f"Could not compute FM from sieve table: {e}")
        raise ValueError(f"FM not found and cannot be computed from {path}")


def compute_fm_from_sieve_table(sheet, file_path: str) -> float:
    """Compute FM from sieve data table.

    Standard sieve sizes: 3/8", No.4, No.8, No.16, No.30, No.50, No.100, No.200
    FM = (sum of cumulative % retained) / 100

    Args:
        sheet: OpenPyXL worksheet object
        file_path: Path to the Excel file

    Returns:
        Computed Fineness Modulus

    Raises:
        ValueError: If sieve data cannot be found or parsed
    """
    pd = _ensure_pandas()
    df = pd.read_excel(file_path, sheet_name=sheet.title, header=None)

    # Look for sieve sizes row
    sieve_sizes = [
        "3/8",
        "No.4",
        "No.8",
        "No.16",
        "No.30",
        "No.50",
        "No.100",
        "No.200",
    ]

    cumulative_retained = []

    # Search for sieve headers and extract cumulative % retained
    for idx, row in df.iterrows():
        row_str = " ".join([str(x).lower() for x in row if pd.notna(x)])

        # Check if this row contains sieve sizes
        found_sieves = []
        for sieve in sieve_sizes:
            if sieve.lower().replace(".", "") in row_str:
                found_sieves.append(sieve)

        if len(found_sieves) >= 4:  # Found at least 4 sieve sizes
            # Look for cumulative % retained in next few rows
            for next_idx in range(idx + 1, min(idx + 5, len(df))):
                next_row = df.iloc[next_idx]

                # Check if this row contains cumulative retained data
                values = []
                for val in next_row:
                    numeric = _coerce_numeric(val)
                    if numeric is not None and 0 <= numeric <= 100:
                        values.append(float(numeric))

                if len(values) >= len(found_sieves):
                    # Found cumulative retained values
                    cumulative_retained = values[: len(found_sieves)]
                    break

            if cumulative_retained:
                break

    if not cumulative_retained:
        raise ValueError("Cannot find sieve table data in sheet")

    # Calculate FM
    fm = sum(cumulative_retained) / 100.0
    return round(fm, 2)


def extract_location_from_xlsx(
    path: str,
    workbook=None,
) -> Tuple[float, float]:
    """Extract location coordinates from Excel file.

    Looks for a cell labeled "Location" and parses its value.

    Args:
        path: Path to Excel file

    Returns:
        Tuple of (latitude, longitude)

    Raises:
        ValueError: If Location cell not found or unparseable
    """
    wb = workbook or _load_workbook(str(Path(path)))
    sh = wb.active

    # Search for "Location" cell
    for row in sh.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue

            cell_text = cell.value.lower()
            if "location" not in cell_text:
                continue

            # Skip labels like "Lab Location" that rarely include coordinates
            if re.search(r"\blab\s+location\b", cell_text):
                continue

            # Some layouts merge cells or leave blanks between label and value.
            # Scan a few cells to the right; ignore unparseable values.
            for offset in range(1, 6):
                next_cell = sh.cell(
                    row=cell.row,
                    column=cell.column + offset,
                )
                if next_cell.value in (None, ""):
                    continue
                try:
                    return parse_location_cell(str(next_cell.value))
                except ValueError:
                    continue

            # Occasionally the location value lives in the same cell.
            value_match = re.search(
                r"location[:\s]+(.+)",
                cell.value,
                flags=re.IGNORECASE,
            )
            if value_match:
                try:
                    return parse_location_cell(value_match.group(1).strip())
                except ValueError:
                    continue

    # Alternative: Look in a specific cell (e.g., B1, A2, etc.)
    common_location_cells = ["B1", "A2", "B2", "C1"]
    for cell_ref in common_location_cells:
        try:
            cell = sh[cell_ref]
            if cell.value:
                try:
                    return parse_location_cell(str(cell.value))
                except ValueError:
                    continue
        except Exception:
            continue

    # Fallback: scan entire sheet for coordinate-like values
    for row in sh.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str):
                try:
                    return parse_location_cell(value)
                except ValueError:
                    continue

    filename = Path(path).name
    raise ValueError(f"Location cell not found in {filename}")


def parse_file(file_path: str, hole_id: Optional[str] = None) -> dict:
    """Parse a gradation report Excel file to extract all metadata.

    Args:
        file_path: Path to Excel file
        hole_id: Optional hole ID (if not in filename)

    Returns:
        Dictionary with hole_id, start_ft, end_ft, latitude, longitude,
        fm_value, and accumulated warnings

    Raises:
        ValueError: If required data cannot be extracted
    """
    path = Path(file_path)
    filename = path.name
    path_str = str(path)
    workbook = _load_workbook(path_str)

    warnings: list[str] = []

    inferred_from_title = parse_hole_id_from_title(filename)
    inferred_from_sheet = parse_hole_id_from_sheet(workbook)

    resolved_hole_id = hole_id or inferred_from_title or inferred_from_sheet
    if not resolved_hole_id:
        raise ValueError(f"Cannot determine hole ID for {filename}")

    resolved_hole_id = resolved_hole_id.upper()

    if (
        hole_id
        and inferred_from_title
        and inferred_from_title.upper() != hole_id.upper()
    ):
        warning_template = (
            "Hole ID mismatch for {file}: folder '{folder}' vs title '{title}'"
        )
        warnings.append(
            warning_template.format(
                file=filename,
                folder=hole_id,
                title=inferred_from_title,
            )
        )
    if (
        hole_id
        and inferred_from_sheet
        and inferred_from_sheet.upper() != hole_id.upper()
    ):
        warning_template = (
            "Hole ID mismatch for {file}: folder '{folder}' vs sheet '{sheet}'"
        )
        warnings.append(
            warning_template.format(
                file=filename,
                folder=hole_id,
                sheet=inferred_from_sheet,
            )
        )
    if not hole_id and inferred_from_title and inferred_from_sheet:
        if inferred_from_title != inferred_from_sheet:
            conflict_template = (
                "Hole ID conflict in {file}: title '{title}' vs "
                "sheet '{sheet}'"
            )
            warnings.append(
                conflict_template.format(
                    file=filename,
                    title=inferred_from_title,
                    sheet=inferred_from_sheet,
                )
            )

    # Extract interval (filename is the source of truth)
    interval_from_title = parse_interval_from_title(filename)
    if not interval_from_title:
        raise ValueError(f"Cannot extract interval for {filename}")

    interval_from_sheet = parse_interval_from_sheet(path_str, workbook)
    if interval_from_sheet and interval_from_sheet != interval_from_title:
        warnings.append(
            (
                "Interval mismatch in {file}: title {title_interval} vs "
                "sheet {sheet_interval}; ignoring sheet"
            ).format(
                file=filename,
                title_interval=interval_from_title,
                sheet_interval=interval_from_sheet,
            )
        )

    start_ft, end_ft = interval_from_title

    # Extract location
    try:
        lat, lon = extract_location_from_xlsx(path_str, workbook)
    except Exception as e:
        logger.error(f"Error extracting location from {filename}: {e}")
        raise ValueError(f"Location extraction failed for {filename}: {e}")

    # Extract FM
    try:
        fm_value = extract_fm_from_xlsx(path_str, workbook)
    except Exception as e:
        logger.error(f"Error extracting FM from {filename}: {e}")
        raise ValueError(f"FM extraction failed for {filename}: {e}")

    return {
        "hole_id": resolved_hole_id,
        "start_ft": start_ft,
        "end_ft": end_ft,
        "latitude": lat,
        "longitude": lon,
        "fm_value": fm_value,
        "filename": filename,
        "warnings": warnings,
    }
