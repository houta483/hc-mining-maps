"""Create stub Excel files for testing sieve analysis parsing."""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Color scheme for headers
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CELL_FONT = Font(size=10)


def create_sieve_analysis_file(
    output_path: str,
    hole_id: str,
    start_ft: int,
    end_ft: int,
    latitude: float,
    longitude: float,
    fm_value: float,
):
    """Create a stub sieve analysis Excel file.

    Args:
        output_path: Path to save the Excel file
        hole_id: Hole identifier (e.g., "T2", "T3")
        start_ft: Starting depth in feet
        end_ft: Ending depth in feet
        latitude: Latitude coordinate
        longitude: Longitude coordinate
        fm_value: Fineness Modulus value
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sieve Analysis"

    # Header
    header_row = 1
    headers = [
        "Sample ID",
        "Depth",
        "Sieve Size",
        "Weight Retained",
        "Cumulative % Passing",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample data - Standard sieve sizes (in mm)
    sieve_sizes = ["4.75", "2.36", "1.18", "0.600", "0.300", "0.150", "0.075", "Pan"]
    sample_id = f"{hole_id}-{start_ft}-{end_ft}"
    depth = f"{start_ft}.0-{end_ft}.0"

    # Generate realistic weight retained data
    import random

    random.seed(hash(f"{hole_id}{start_ft}{end_ft}") % 1000)  # Deterministic randomness

    weight_retained = []
    cumulative_passing = [100.0]

    for i in range(len(sieve_sizes) - 1):  # Exclude Pan
        wt = round(random.uniform(0.5, 25.0), 1)
        weight_retained.append(wt)
        remaining = cumulative_passing[-1] - (wt / 100.0 * cumulative_passing[-1])
        cumulative_passing.append(round(remaining, 1))

    # Pan gets remainder
    weight_retained.append(round(100.0 - sum(weight_retained[:-1]), 1))
    cumulative_passing[-1] = 0.0

    # Write data rows
    for idx, (sieve, wt, cum) in enumerate(
        zip(sieve_sizes, weight_retained, cumulative_passing)
    ):
        row = header_row + 1 + idx
        ws.cell(row=row, column=1, value=sample_id).font = CELL_FONT
        ws.cell(row=row, column=2, value=depth).font = CELL_FONT
        ws.cell(
            row=row, column=3, value=f"{sieve} mm" if sieve != "Pan" else "Pan"
        ).font = CELL_FONT
        ws.cell(row=row, column=4, value=wt).font = CELL_FONT
        ws.cell(row=row, column=5, value=cum).font = CELL_FONT

    # Add Total row
    total_row = header_row + len(sieve_sizes) + 1
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(weight_retained)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value="Sum").font = Font(bold=True)

    # Add Location data (in row near top, typical pattern in these files)
    location_row = total_row + 3
    ws.cell(row=location_row, column=1, value="Location:")
    ws.cell(row=location_row, column=2, value=f"{latitude}, {longitude}")

    # Add Fineness Modulus (FM) value
    fm_row = location_row + 1
    ws.cell(row=fm_row, column=1, value="Fineness Modulus:")
    ws.cell(row=fm_row, column=2, value=fm_value)

    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22

    # Save file
    wb.save(output_path)
    print(f"Created: {output_path}")


def main():
    """Create stub data files matching the expected structure."""
    base_dir = Path("test_data")
    base_dir.mkdir(exist_ok=True)

    # Base coordinates: 32°29'04.82" N 96°25'58.72" W
    # Converted to decimal: 32.484672, -96.432978
    base_lat = 32.484672
    base_lon = -96.432978

    # Small variations (within ~200 feet = ~0.0005 degrees)
    # Each hole will have slightly different coordinates
    import random

    random.seed(42)  # Deterministic randomness

    def get_hole_coords(hole_num):
        """Get coordinates for a hole, slightly offset from base."""
        # Use hole_num as seed so same hole always gets same coordinates
        random.seed(hole_num)
        # Add small random offset up to 0.0005 degrees (~200 feet)
        lat_offset = random.uniform(-0.0005, 0.0005)
        lon_offset = random.uniform(-0.0005, 0.0005)
        return base_lat + lat_offset, base_lon + lon_offset

    # UP-B mine area with T2, T3, T4, T5, T6 hole folders
    # All coordinates are near 32.484672, -96.432978 (within ~200 feet)
    # Each hole has consistent coordinates for all its intervals
    mine_areas = {
        "UP-B": {
            "T2": [
                ("T2", 5, 10, *get_hole_coords(2), 2.85),
            ],
            "T3": [
                ("T3", 5, 10, *get_hole_coords(3), 2.92),
                ("T3", 10, 15, *get_hole_coords(3), 3.15),
                ("T3", 15, 20, *get_hole_coords(3), 2.78),
            ],
            "T4": [
                ("T4", 5, 10, *get_hole_coords(4), 3.22),
                ("T4", 10, 15, *get_hole_coords(4), 3.45),
                ("T4", 15, 20, *get_hole_coords(4), 2.95),
            ],
            "T5": [
                ("T5", 5, 10, *get_hole_coords(5), 2.68),
                ("T5", 10, 15, *get_hole_coords(5), 3.08),
                ("T5", 15, 20, *get_hole_coords(5), 2.82),
            ],
            "T6": [
                ("T6", 5, 10, *get_hole_coords(6), 3.11),
            ],
        }
    }

    for mine_area, holes in mine_areas.items():
        mine_dir = base_dir / mine_area
        mine_dir.mkdir(exist_ok=True)

        for hole_id, intervals in holes.items():
            hole_dir = mine_dir / hole_id
            hole_dir.mkdir(exist_ok=True)

            for hole_id_val, start, end, lat, lon, fm in intervals:
                filename = f"QC Gradation Report {hole_id_val} {start}_{end}.xlsx"
                filepath = hole_dir / filename

                create_sieve_analysis_file(
                    str(filepath),
                    hole_id_val,
                    start,
                    end,
                    lat,
                    lon,
                    fm,
                )

    print(f"\n✓ Created stub data in {base_dir.absolute()}")
    print(f"\nTo use this data, set USE_LOCAL_DATA=true in .env or run with:")
    print(f"  USE_LOCAL_DATA=true docker-compose up")


if __name__ == "__main__":
    main()
